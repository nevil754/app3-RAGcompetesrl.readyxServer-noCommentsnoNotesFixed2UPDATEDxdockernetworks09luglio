from __future__ import annotations
import mimetypes
from pathlib import Path
from typing import Any
from loguru import logger
from app.core.settings import get_settings

settings = get_settings()

class ParsedDocument:
    def __init__(
        self,
        text: str,
        pages: list[str],
        tables: list[dict],
        metadata: dict[str, Any],
        page_count: int,
    ):
        self.text = text
        self.pages = pages
        self.tables = tables
        self.metadata = metadata
        self.page_count = page_count

def parse_document(file_path: str) -> ParsedDocument:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File non trovato: {file_path}")
    suffix = path.suffix.lower()
    mime_type = mimetypes.guess_type(file_path)[0] or ""
    logger.info(f"Parsing documento: {path.name} ({suffix})")
    if suffix in {".pdf", ".docx", ".pptx"} and settings.ingestion_prefer_docling:
        try:
            return _parse_with_docling(file_path)
        except Exception as e:
            logger.warning(f"Docling fallito ({e}), fallback su unstructured")
            return _parse_with_unstructured(file_path)
    elif suffix in {".xlsx", ".xls"}:
        return _parse_excel(file_path)
    elif suffix in {".txt", ".md"}:
        return _parse_text(file_path)
    else:
        return _parse_with_unstructured(file_path)

def _parse_with_docling(file_path: str) -> ParsedDocument:
    from docling.document_converter import DocumentConverter
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    pipeline_options = PdfPipelineOptions()
    pipeline_options.do_table_structure = settings.ingestion_extract_tables
    pipeline_options.do_ocr = False
    converter = DocumentConverter()
    result = converter.convert(file_path)
    doc = result.document

    full_text = doc.export_to_markdown()

    page_texts: dict[int, list[str]] = {}
    for item, _level in doc.iterate_items():
        prov_list = getattr(item, "prov", None)
        if not prov_list:
            continue
        page_no = prov_list[0].page_no
        content = getattr(item, "text", None) or ""
        if content:
            page_texts.setdefault(page_no, []).append(content)

    pages = [ "\n".join(page_texts.get(p, [])) for p in range(1, len(doc.pages) + 1) ]
    tables = []
    if settings.ingestion_extract_tables:
        for table in doc.tables:
            tables.append({
                "page": table.prov[0].page_no if table.prov else None,
                "markdown": table.export_to_markdown(),
            })
    metadata = {
        "parser": "docling",
        "page_count": len(doc.pages),
        "table_count": len(tables),
        "has_tables": len(tables) > 0,
    }
    logger.debug(
        "Docling parsing completato",
        pages=len(pages),
        tables=len(tables),
        chars=len(full_text),
    )
    return ParsedDocument(
        text=full_text,
        pages=pages,
        tables=tables,
        metadata=metadata,
        page_count=len(pages),
    )

def _parse_excel(file_path: str) -> ParsedDocument:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets_text: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(" | ".join(str(c or "") for c in row))
        if rows:
            sheet_md = f"## Foglio: {sheet_name}\n\n" + "\n".join(rows)
            sheets_text.append(sheet_md)
    full_text = "\n\n".join(sheets_text)
    return ParsedDocument(
        text=full_text,
        pages=[full_text],
        tables=[],
        metadata={"parser": "openpyxl", "sheets": wb.sheetnames},
        page_count=1,
    )

def _parse_text(file_path: str) -> ParsedDocument:
    with open(file_path, encoding="utf-8", errors="replace") as f:
        text = f.read()
    return ParsedDocument(
        text=text,
        pages=[text],
        tables=[],
        metadata={"parser": "text"},
        page_count=1,
    )

def _parse_with_unstructured(file_path: str) -> ParsedDocument:
    from unstructured.partition.auto import partition
    elements = partition(
        filename=file_path,
        include_page_breaks=True,
        strategy="fast",
    )
    pages: list[str] = []
    current_page: list[str] = []
    tables: list[dict] = []
    for el in elements:
        el_type = type(el).__name__
        text = str(el)
        if el_type == "PageBreak":
            pages.append("\n".join(current_page))
            current_page = []
        elif el_type == "Table":
            tables.append({"text": text, "page": len(pages) + 1})
            current_page.append(text)
        else:
            current_page.append(text)
    if current_page:
        pages.append("\n".join(current_page))
    full_text = "\n\n".join(pages)
    return ParsedDocument(
        text=full_text,
        pages=pages,
        tables=tables,
        metadata={"parser": "unstructured", "page_count": len(pages)},
        page_count=len(pages),
    )
